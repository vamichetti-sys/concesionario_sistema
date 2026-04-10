from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db.models import Sum
from django.http import HttpResponse
from datetime import date
from decimal import Decimal

from .models import FacturaRegistrada, CompraRegistrada
from .forms import FacturaRegistradaForm, CompraRegistradaForm

# EXCEL
from openpyxl import Workbook

# PDF
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors


# ==========================================================
# IDENTIDAD VISUAL
# ==========================================================
COLOR_AZUL = colors.HexColor("#002855")
COLOR_NARANJA = colors.HexColor("#FF6C1A")
COLOR_GRIS = colors.HexColor("#F4F6F8")


# ==========================================================
# LISTA FACTURACIÓN
# ==========================================================
@login_required
def lista_facturacion(request):
    hoy = date.today()
    mes = int(request.GET.get("mes", hoy.month))
    anio = int(request.GET.get("anio", hoy.year))

    # Listado completo (todas las facturas válidas, sin filtrar por mes)
    facturas = (
        FacturaRegistrada.objects
        .filter(estado="valida")
        .order_by("-fecha")
    )

    # Totales del mes filtrado
    total_mes = (
        FacturaRegistrada.objects
        .filter(
            estado="valida",
            fecha__year=anio,
            fecha__month=mes,
        )
        .aggregate(total=Sum("monto"))["total"] or Decimal("0")
    )

    # Totales del año filtrado
    total_anio = (
        FacturaRegistrada.objects
        .filter(
            estado="valida",
            fecha__year=anio,
        )
        .aggregate(total=Sum("monto"))["total"] or Decimal("0")
    )

    MESES = [
        "", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
        "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
    ]

    return render(
        request,
        "facturacion/lista.html",
        {
            "page_title": "Facturación",
            "facturas": facturas,
            "mes": mes,
            "anio": anio,
            "mes_nombre": MESES[mes] if 1 <= mes <= 12 else "",
            "meses_choices": list(enumerate(MESES))[1:],
            "total_mes": total_mes,
            "total_anio": total_anio,
        }
    )


# ==========================================================
# CREAR FACTURA
# ==========================================================
@login_required
def crear_factura(request):
    if request.method == "POST":
        form = FacturaRegistradaForm(request.POST)
        if form.is_valid():
            factura = form.save(commit=False)
            # 🔒 Seguridad extra: recalcular IVA
            factura.calcular_iva()
            factura.save()
            return redirect("facturacion:lista")
    else:
        form = FacturaRegistradaForm()

    return render(
        request,
        "facturacion/crear.html",
        {
            "page_title": "Registrar Factura",
            "form": form,
        }
    )


# ==========================================================
# EXPORTAR EXCEL MENSUAL
# ==========================================================
@login_required
def exportar_excel_mensual(request):
    hoy = date.today()
    mes = int(request.GET.get("mes", hoy.month))
    anio = int(request.GET.get("anio", hoy.year))

    facturas = FacturaRegistrada.objects.filter(
        estado="valida",
        fecha__year=anio,
        fecha__month=mes,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Facturación Mensual"

    ws.append(["Número", "Fecha", "Monto total", "Venta"])

    for f in facturas:
        ws.append([
            f.numero,
            f.fecha.strftime("%d/%m/%Y"),
            float(f.monto),
            f.venta.id if f.venta else ""
        ])

    total = facturas.aggregate(total=Sum("monto"))["total"] or 0
    ws.append([])
    ws.append(["TOTAL", "", float(total), ""])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="facturacion_{hoy.month}_{hoy.year}.xlsx"'
    )

    wb.save(response)
    return response


# ==========================================================
# EXPORTAR EXCEL ANUAL
# ==========================================================
@login_required
def exportar_excel_anual(request):
    hoy = date.today()
    anio = int(request.GET.get("anio", hoy.year))

    facturas = FacturaRegistrada.objects.filter(
        estado="valida",
        fecha__year=anio,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Facturación Anual"

    ws.append(["Número", "Fecha", "Monto total", "Venta"])

    for f in facturas:
        ws.append([
            f.numero,
            f.fecha.strftime("%d/%m/%Y"),
            float(f.monto),
            f.venta.id if f.venta else ""
        ])

    total = facturas.aggregate(total=Sum("monto"))["total"] or 0
    ws.append([])
    ws.append(["TOTAL ANUAL", "", float(total), ""])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="facturacion_anual_{anio}.xlsx"'
    )

    wb.save(response)
    return response


# ==========================================================
# EXPORTAR PDF MENSUAL (CON IVA DISCRIMINADO)
# ==========================================================
@login_required
def exportar_pdf_mensual(request):
    hoy = date.today()
    mes = int(request.GET.get("mes", hoy.month))
    anio = int(request.GET.get("anio", hoy.year))

    facturas = FacturaRegistrada.objects.filter(
        estado="valida",
        fecha__year=anio,
        fecha__month=mes,
    )

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = (
        f'inline; filename="facturacion_{mes}_{anio}.pdf"'
    )

    doc = SimpleDocTemplate(
        response,
        pagesize=A4,
        rightMargin=40,
        leftMargin=40,
        topMargin=40,
        bottomMargin=40
    )

    styles = getSampleStyleSheet()
    elementos = []

    MESES_NOMBRE = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
                    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]

    header = Table(
        [[
            Paragraph(
                "<b>AMICHETTI AUTOMOTORES</b><br/>"
                f"Facturación Mensual – {MESES_NOMBRE[mes]} {anio}",
                ParagraphStyle(
                    "h1",
                    fontSize=14,
                    textColor=colors.white
                )
            ),
            Paragraph(
                f"Fecha emisión<br/>{hoy.strftime('%d/%m/%Y')}",
                ParagraphStyle(
                    "h2",
                    fontSize=10,
                    textColor=colors.white,
                    alignment=2
                )
            )
        ]],
        colWidths=[340, 180]
    )

    header.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), COLOR_AZUL),
        ("PADDING", (0, 0), (-1, -1), 14),
    ]))

    elementos.append(header)
    elementos.append(Spacer(1, 20))

    data = [
        ["N° Factura", "Fecha", "Neto", "IVA", "Total"]
    ]

    for f in facturas:
        neto = f.monto_neto if f.monto_neto is not None else f.monto
        iva = f.monto_iva if f.monto_iva is not None else Decimal("0.00")
        total = f.monto

        data.append([
            f.numero,
            f.fecha.strftime("%d/%m/%Y"),
            f"$ {neto:,.2f}",
            f"$ {iva:,.2f}",
            f"$ {total:,.2f}",
        ])

    tabla = Table(data, colWidths=[90, 80, 110, 110, 110])
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), COLOR_GRIS),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (2, 1), (-1, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ("PADDING", (0, 0), (-1, -1), 8),
    ]))

    elementos.append(tabla)
    elementos.append(Spacer(1, 16))

    total_general = facturas.aggregate(total=Sum("monto"))["total"] or 0

    total_box = Table(
        [["TOTAL FACTURADO", f"$ {total_general:,.2f}"]],
        colWidths=[390, 110]
    )

    total_box.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 13),
        ("TEXTCOLOR", (1, 0), (1, 0), COLOR_NARANJA),
        ("ALIGN", (1, 0), (1, 0), "RIGHT"),
        ("LINEABOVE", (0, 0), (-1, 0), 1, colors.lightgrey),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
    ]))

    elementos.append(total_box)
    elementos.append(Spacer(1, 30))

    elementos.append(
        Paragraph(
            "Amichetti Automotores · Rojas, Buenos Aires",
            ParagraphStyle(
                "f",
                fontSize=8,
                textColor=colors.grey,
                alignment=1
            )
        )
    )

    doc.build(elementos)
    return response


# ==========================================================
# EXPORTAR PDF ANUAL (CON IVA DISCRIMINADO)
# ==========================================================
@login_required
def exportar_pdf_anual(request):
    hoy = date.today()
    anio = int(request.GET.get("anio", hoy.year))

    facturas = FacturaRegistrada.objects.filter(
        estado="valida",
        fecha__year=anio,
    )

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = (
        f'inline; filename="facturacion_anual_{anio}.pdf"'
    )

    doc = SimpleDocTemplate(
        response,
        pagesize=A4,
        rightMargin=40,
        leftMargin=40,
        topMargin=40,
        bottomMargin=40
    )

    styles = getSampleStyleSheet()
    elementos = []

    header = Table(
        [[
            Paragraph(
                "<b>AMICHETTI AUTOMOTORES</b><br/>"
                f"Facturación Anual – {anio}",
                ParagraphStyle(
                    "h1",
                    fontSize=14,
                    textColor=colors.white
                )
            ),
            Paragraph(
                f"Fecha emisión<br/>{hoy.strftime('%d/%m/%Y')}",
                ParagraphStyle(
                    "h2",
                    fontSize=10,
                    textColor=colors.white,
                    alignment=2
                )
            )
        ]],
        colWidths=[340, 180]
    )

    header.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), COLOR_AZUL),
        ("PADDING", (0, 0), (-1, -1), 14),
    ]))

    elementos.append(header)
    elementos.append(Spacer(1, 20))

    data = [
        ["N° Factura", "Fecha", "Neto", "IVA", "Total"]
    ]

    for f in facturas:
        neto = f.monto_neto if f.monto_neto is not None else f.monto
        iva = f.monto_iva if f.monto_iva is not None else Decimal("0.00")
        total = f.monto

        data.append([
            f.numero,
            f.fecha.strftime("%d/%m/%Y"),
            f"$ {neto:,.2f}",
            f"$ {iva:,.2f}",
            f"$ {total:,.2f}",
        ])

    tabla = Table(data, colWidths=[90, 80, 110, 110, 110])
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), COLOR_GRIS),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (2, 1), (-1, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ("PADDING", (0, 0), (-1, -1), 8),
    ]))

    elementos.append(tabla)
    elementos.append(Spacer(1, 16))

    total_general = facturas.aggregate(total=Sum("monto"))["total"] or 0

    total_box = Table(
        [["TOTAL ANUAL", f"$ {total_general:,.2f}"]],
        colWidths=[390, 110]
    )

    total_box.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 13),
        ("TEXTCOLOR", (1, 0), (1, 0), COLOR_NARANJA),
        ("ALIGN", (1, 0), (1, 0), "RIGHT"),
        ("LINEABOVE", (0, 0), (-1, 0), 1, colors.lightgrey),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
    ]))

    elementos.append(total_box)
    elementos.append(Spacer(1, 30))

    elementos.append(
        Paragraph(
            "Amichetti Automotores · Rojas, Buenos Aires",
            ParagraphStyle(
                "f",
                fontSize=8,
                textColor=colors.grey,
                alignment=1
            )
        )
    )

    doc.build(elementos)
    return response

# ==========================================================
# ELIMINAR FACTURA
# ==========================================================
@login_required
def eliminar_factura(request, pk):
    factura = get_object_or_404(FacturaRegistrada, pk=pk)

    if request.method == "POST":
        numero = factura.numero
        factura.delete()
        messages.success(request, f"Factura #{numero} eliminada.")
        return redirect("facturacion:lista")

    return render(request, "facturacion/eliminar.html", {"factura": factura})


# ==========================================================
# COMPRAS - LISTA
# ==========================================================
@login_required
def lista_compras(request):
    hoy = date.today()
    mes = int(request.GET.get("mes", hoy.month))
    anio = int(request.GET.get("anio", hoy.year))

    # Listado completo (todas las compras, sin filtrar por mes)
    compras = CompraRegistrada.objects.all().order_by("-fecha")

    # Totales del mes filtrado
    compras_mes = CompraRegistrada.objects.filter(
        fecha__year=anio, fecha__month=mes,
    )
    total_mes = compras_mes.aggregate(t=Sum("monto"))["t"] or Decimal("0")
    total_iva_mes = compras_mes.aggregate(t=Sum("monto_iva"))["t"] or Decimal("0")

    # Totales del año filtrado
    compras_anio = CompraRegistrada.objects.filter(fecha__year=anio)
    total_anio = compras_anio.aggregate(t=Sum("monto"))["t"] or Decimal("0")

    MESES = [
        "", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
        "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
    ]

    return render(request, "facturacion/compras.html", {
        "compras": compras,
        "mes": mes,
        "anio": anio,
        "mes_nombre": MESES[mes] if 1 <= mes <= 12 else "",
        "total_mes": total_mes,
        "total_iva_mes": total_iva_mes,
        "total_anio": total_anio,
        "meses_choices": list(enumerate(MESES))[1:],
    })


# ==========================================================
# COMPRAS - CREAR
# ==========================================================
@login_required
def crear_compra(request):
    if request.method == "POST":
        form = CompraRegistradaForm(request.POST)
        if form.is_valid():
            compra = form.save(commit=False)
            compra.calcular_iva()
            compra.save()
            messages.success(request, f"Compra #{compra.numero} registrada.")
            return redirect("facturacion:compras")
        else:
            messages.error(request, f"Error al guardar: {form.errors.as_text()}")
    else:
        form = CompraRegistradaForm()

    return render(request, "facturacion/crear_compra.html", {
        "form": form,
        "titulo": "Registrar Compra",
    })


# ==========================================================
# COMPRAS - ELIMINAR
# ==========================================================
@login_required
def eliminar_compra(request, pk):
    compra = get_object_or_404(CompraRegistrada, pk=pk)

    if request.method == "POST":
        numero = compra.numero
        compra.delete()
        messages.success(request, f"Compra #{numero} eliminada.")
        return redirect("facturacion:compras")

    return render(request, "facturacion/eliminar_compra.html", {"compra": compra})


# ==========================================================
# COMPRAS - EXPORTAR PDF MENSUAL
# ==========================================================
@login_required
def compras_pdf_mensual(request):
    hoy = date.today()
    mes = int(request.GET.get("mes", hoy.month))
    anio = int(request.GET.get("anio", hoy.year))

    MESES = [
        "", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
        "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
    ]

    compras = CompraRegistrada.objects.filter(
        fecha__year=anio, fecha__month=mes,
    ).order_by("-fecha")

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = (
        f'inline; filename="compras_{mes}_{anio}.pdf"'
    )

    doc = SimpleDocTemplate(response, pagesize=A4, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    elementos = []

    header = Table(
        [[
            Paragraph(
                "<b>AMICHETTI AUTOMOTORES</b><br/>"
                f"Compras – {MESES[mes]} {anio}",
                ParagraphStyle("h1", fontSize=14, textColor=colors.white)
            ),
            Paragraph(
                f"Fecha emisión<br/>{hoy.strftime('%d/%m/%Y')}",
                ParagraphStyle("h2", fontSize=10, textColor=colors.white, alignment=2)
            )
        ]],
        colWidths=[340, 180]
    )
    header.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), COLOR_AZUL),
        ("PADDING", (0, 0), (-1, -1), 14),
    ]))
    elementos.append(header)
    elementos.append(Spacer(1, 20))

    data = [["N° Factura", "Proveedor", "Fecha", "Neto", "IVA", "Total"]]
    for c in compras:
        neto = c.monto_neto or Decimal("0")
        iva = c.monto_iva or Decimal("0")
        total = c.monto or Decimal("0")
        data.append([
            c.numero,
            (c.proveedor or "-")[:20],
            c.fecha.strftime("%d/%m/%Y"),
            f"$ {neto:,.2f}",
            f"$ {iva:,.2f}",
            f"$ {total:,.2f}",
        ])

    tabla = Table(data, colWidths=[80, 90, 70, 90, 90, 90])
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), COLOR_GRIS),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (3, 1), (-1, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ("PADDING", (0, 0), (-1, -1), 8),
    ]))
    elementos.append(tabla)
    elementos.append(Spacer(1, 16))

    total_general = compras.aggregate(total=Sum("monto"))["total"] or 0
    total_box = Table([["TOTAL COMPRAS", f"$ {total_general:,.2f}"]], colWidths=[390, 110])
    total_box.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 13),
        ("TEXTCOLOR", (1, 0), (1, 0), COLOR_NARANJA),
        ("ALIGN", (1, 0), (1, 0), "RIGHT"),
        ("LINEABOVE", (0, 0), (-1, 0), 1, colors.lightgrey),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
    ]))
    elementos.append(total_box)
    elementos.append(Spacer(1, 30))
    elementos.append(Paragraph("Amichetti Automotores · Rojas, Buenos Aires", ParagraphStyle("f", fontSize=8, textColor=colors.grey, alignment=1)))

    doc.build(elementos)
    return response


# ==========================================================
# COMPRAS - EXPORTAR PDF ANUAL
# ==========================================================
@login_required
def compras_pdf_anual(request):
    hoy = date.today()
    anio = int(request.GET.get("anio", hoy.year))

    compras = CompraRegistrada.objects.filter(fecha__year=anio).order_by("-fecha")

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = (
        f'inline; filename="compras_anual_{anio}.pdf"'
    )

    doc = SimpleDocTemplate(response, pagesize=A4, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    elementos = []

    header = Table(
        [[
            Paragraph(
                f"<b>AMICHETTI AUTOMOTORES</b><br/>Compras Anual – {anio}",
                ParagraphStyle("h1", fontSize=14, textColor=colors.white)
            ),
            Paragraph(
                f"Fecha emisión<br/>{hoy.strftime('%d/%m/%Y')}",
                ParagraphStyle("h2", fontSize=10, textColor=colors.white, alignment=2)
            )
        ]],
        colWidths=[340, 180]
    )
    header.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), COLOR_AZUL),
        ("PADDING", (0, 0), (-1, -1), 14),
    ]))
    elementos.append(header)
    elementos.append(Spacer(1, 20))

    data = [["N° Factura", "Proveedor", "Fecha", "Neto", "IVA", "Total"]]
    for c in compras:
        neto = c.monto_neto or Decimal("0")
        iva = c.monto_iva or Decimal("0")
        total = c.monto or Decimal("0")
        data.append([
            c.numero,
            (c.proveedor or "-")[:20],
            c.fecha.strftime("%d/%m/%Y"),
            f"$ {neto:,.2f}",
            f"$ {iva:,.2f}",
            f"$ {total:,.2f}",
        ])

    tabla = Table(data, colWidths=[80, 90, 70, 90, 90, 90])
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), COLOR_GRIS),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (3, 1), (-1, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ("PADDING", (0, 0), (-1, -1), 8),
    ]))
    elementos.append(tabla)
    elementos.append(Spacer(1, 16))

    total_general = compras.aggregate(total=Sum("monto"))["total"] or 0
    total_box = Table([["TOTAL ANUAL COMPRAS", f"$ {total_general:,.2f}"]], colWidths=[390, 110])
    total_box.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 13),
        ("TEXTCOLOR", (1, 0), (1, 0), COLOR_NARANJA),
        ("ALIGN", (1, 0), (1, 0), "RIGHT"),
        ("LINEABOVE", (0, 0), (-1, 0), 1, colors.lightgrey),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
    ]))
    elementos.append(total_box)
    elementos.append(Spacer(1, 30))
    elementos.append(Paragraph("Amichetti Automotores · Rojas, Buenos Aires", ParagraphStyle("f", fontSize=8, textColor=colors.grey, alignment=1)))

    doc.build(elementos)
    return response


# ==========================================================
# COMPRAS - EXPORTAR EXCEL MENSUAL
# ==========================================================
@login_required
def compras_excel_mensual(request):
    hoy = date.today()
    mes = int(request.GET.get("mes", hoy.month))
    anio = int(request.GET.get("anio", hoy.year))

    compras = CompraRegistrada.objects.filter(
        fecha__year=anio, fecha__month=mes,
    ).order_by("-fecha")

    wb = Workbook()
    ws = wb.active
    ws.title = f"Compras {mes}-{anio}"
    ws.append(["N° Factura", "Proveedor", "Fecha", "Neto", "IVA", "Otros Imp.", "Total"])

    for c in compras:
        ws.append([
            c.numero,
            c.proveedor or "-",
            c.fecha.strftime("%d/%m/%Y"),
            float(c.monto_neto or 0),
            float(c.monto_iva or 0),
            float(c.otros_impuestos or 0),
            float(c.monto or 0),
        ])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="compras_{mes}_{anio}.xlsx"'
    wb.save(response)
    return response


# ==========================================================
# COMPRAS - EXPORTAR EXCEL ANUAL
# ==========================================================
@login_required
def compras_excel_anual(request):
    hoy = date.today()
    anio = int(request.GET.get("anio", hoy.year))

    compras = CompraRegistrada.objects.filter(fecha__year=anio).order_by("-fecha")

    wb = Workbook()
    ws = wb.active
    ws.title = f"Compras {anio}"
    ws.append(["N° Factura", "Proveedor", "Fecha", "Neto", "IVA", "Otros Imp.", "Total"])

    for c in compras:
        ws.append([
            c.numero,
            c.proveedor or "-",
            c.fecha.strftime("%d/%m/%Y"),
            float(c.monto_neto or 0),
            float(c.monto_iva or 0),
            float(c.otros_impuestos or 0),
            float(c.monto or 0),
        ])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="compras_anual_{anio}.xlsx"'
    wb.save(response)
    return response


# ==========================================================
# IVA - POSICION MENSUAL
# ==========================================================
@login_required
def posicion_iva(request):
    hoy = date.today()
    mes = int(request.GET.get("mes", hoy.month))
    anio = int(request.GET.get("anio", hoy.year))

    facturas = FacturaRegistrada.objects.filter(
        estado="valida", fecha__year=anio, fecha__month=mes,
    )
    iva_debito = facturas.aggregate(t=Sum("monto_iva"))["t"] or Decimal("0")
    neto_ventas = facturas.aggregate(t=Sum("monto_neto"))["t"] or Decimal("0")
    total_ventas = facturas.aggregate(t=Sum("monto"))["t"] or Decimal("0")

    compras = CompraRegistrada.objects.filter(
        fecha__year=anio, fecha__month=mes,
    )
    iva_credito = compras.aggregate(t=Sum("monto_iva"))["t"] or Decimal("0")
    neto_compras = compras.aggregate(t=Sum("monto_neto"))["t"] or Decimal("0")
    total_compras = compras.aggregate(t=Sum("monto"))["t"] or Decimal("0")

    saldo_iva = iva_debito - iva_credito

    # IVA a favor acumulado de meses anteriores
    iva_debito_anterior = FacturaRegistrada.objects.filter(
        estado="valida", fecha__year=anio, fecha__month__lt=mes,
    ).aggregate(t=Sum("monto_iva"))["t"] or Decimal("0")

    iva_credito_anterior = CompraRegistrada.objects.filter(
        fecha__year=anio, fecha__month__lt=mes,
    ).aggregate(t=Sum("monto_iva"))["t"] or Decimal("0")

    saldo_anterior = iva_debito_anterior - iva_credito_anterior
    iva_a_favor_acumulado = min(saldo_anterior, Decimal("0"))  # negativo = a favor

    saldo_final = saldo_iva + iva_a_favor_acumulado

    MESES = [
        "", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
        "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
    ]

    return render(request, "facturacion/iva.html", {
        "mes": mes,
        "anio": anio,
        "mes_nombre": MESES[mes] if 1 <= mes <= 12 else "",
        "meses_choices": list(enumerate(MESES))[1:],
        "iva_debito": iva_debito,
        "neto_ventas": neto_ventas,
        "total_ventas": total_ventas,
        "facturas": facturas,
        "facturas_count": facturas.count(),
        "iva_credito": iva_credito,
        "neto_compras": neto_compras,
        "total_compras": total_compras,
        "compras": compras,
        "compras_count": compras.count(),
        "saldo_iva": saldo_iva,
        "iva_a_favor_acumulado": iva_a_favor_acumulado,
        "saldo_final": saldo_final,
    })


# ==========================================================
# IVA - EXPORTAR PDF POSICION MENSUAL
# ==========================================================
@login_required
def iva_pdf(request):
    hoy = date.today()
    mes = int(request.GET.get("mes", hoy.month))
    anio = int(request.GET.get("anio", hoy.year))

    facturas = FacturaRegistrada.objects.filter(
        estado="valida", fecha__year=anio, fecha__month=mes,
    )
    compras = CompraRegistrada.objects.filter(
        fecha__year=anio, fecha__month=mes,
    )

    iva_debito = facturas.aggregate(t=Sum("monto_iva"))["t"] or Decimal("0")
    neto_ventas = facturas.aggregate(t=Sum("monto_neto"))["t"] or Decimal("0")
    total_ventas = facturas.aggregate(t=Sum("monto"))["t"] or Decimal("0")

    iva_credito = compras.aggregate(t=Sum("monto_iva"))["t"] or Decimal("0")
    neto_compras = compras.aggregate(t=Sum("monto_neto"))["t"] or Decimal("0")
    total_compras = compras.aggregate(t=Sum("monto"))["t"] or Decimal("0")

    saldo_iva = iva_debito - iva_credito

    MESES = [
        "", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
        "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
    ]
    mes_nombre = MESES[mes] if 1 <= mes <= 12 else ""

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="iva_{mes}_{anio}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    elementos = []

    # Header
    header = Table(
        [[
            Paragraph(
                f"<b>AMICHETTI AUTOMOTORES</b><br/>Posicion IVA - {mes_nombre} {anio}",
                ParagraphStyle("h1", fontSize=14, textColor=colors.white)
            ),
            Paragraph(
                f"Generado el {hoy.strftime('%d/%m/%Y')}",
                ParagraphStyle("h2", fontSize=10, textColor=colors.white, alignment=2)
            ),
        ]],
        colWidths=[340, 180]
    )
    header.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), COLOR_AZUL),
        ("PADDING", (0, 0), (-1, -1), 14),
    ]))
    elementos.append(header)
    elementos.append(Spacer(1, 20))

    # Ventas
    elementos.append(Paragraph("<b>IVA DEBITO FISCAL (VENTAS)</b>", ParagraphStyle("s", fontSize=11, textColor=COLOR_AZUL)))
    elementos.append(Spacer(1, 8))

    data_v = [["N° Factura", "Fecha", "Neto", "IVA", "Total"]]
    for f in facturas:
        neto = f.monto_neto if f.monto_neto else f.monto
        iva = f.monto_iva if f.monto_iva else Decimal("0")
        data_v.append([f.numero, f.fecha.strftime("%d/%m/%Y"), f"$ {neto:,.2f}", f"$ {iva:,.2f}", f"$ {f.monto:,.2f}"])

    if not facturas.exists():
        data_v.append(["—", "—", "—", "—", "Sin facturas"])

    tabla_v = Table(data_v, colWidths=[90, 80, 110, 110, 110])
    tabla_v.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), COLOR_GRIS),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (2, 1), (-1, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ("PADDING", (0, 0), (-1, -1), 6),
    ]))
    elementos.append(tabla_v)
    elementos.append(Spacer(1, 6))
    elementos.append(Paragraph(f"<b>Total IVA Debito: $ {iva_debito:,.2f}</b>", ParagraphStyle("td", fontSize=10, alignment=2)))
    elementos.append(Spacer(1, 16))

    # Compras
    elementos.append(Paragraph("<b>IVA CREDITO FISCAL (COMPRAS)</b>", ParagraphStyle("s2", fontSize=11, textColor=COLOR_AZUL)))
    elementos.append(Spacer(1, 8))

    data_c = [["N° Factura", "Proveedor", "Neto", "IVA", "Total"]]
    for c in compras:
        neto = c.monto_neto if c.monto_neto else c.monto
        iva = c.monto_iva if c.monto_iva else Decimal("0")
        data_c.append([c.numero, (c.proveedor or "-")[:20], f"$ {neto:,.2f}", f"$ {iva:,.2f}", f"$ {c.monto:,.2f}"])

    if not compras.exists():
        data_c.append(["—", "—", "—", "—", "Sin compras"])

    tabla_c = Table(data_c, colWidths=[90, 100, 100, 100, 110])
    tabla_c.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), COLOR_GRIS),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (2, 1), (-1, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ("PADDING", (0, 0), (-1, -1), 6),
    ]))
    elementos.append(tabla_c)
    elementos.append(Spacer(1, 6))
    elementos.append(Paragraph(f"<b>Total IVA Credito: $ {iva_credito:,.2f}</b>", ParagraphStyle("tc", fontSize=10, alignment=2)))
    elementos.append(Spacer(1, 20))

    # Resultado
    resultado_texto = "A PAGAR" if saldo_iva > 0 else "A FAVOR" if saldo_iva < 0 else "NEUTRO"
    resultado_color = COLOR_NARANJA if saldo_iva > 0 else colors.HexColor("#10b981")

    total_box = Table(
        [[f"SALDO IVA ({resultado_texto})", f"$ {saldo_iva:,.2f}"]],
        colWidths=[390, 110]
    )
    total_box.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 13),
        ("TEXTCOLOR", (1, 0), (1, 0), resultado_color),
        ("ALIGN", (1, 0), (1, 0), "RIGHT"),
        ("LINEABOVE", (0, 0), (-1, 0), 1, colors.lightgrey),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
    ]))
    elementos.append(total_box)
    elementos.append(Spacer(1, 30))

    elementos.append(Paragraph(
        "Amichetti Automotores - Rojas, Buenos Aires",
        ParagraphStyle("footer", fontSize=8, textColor=colors.grey, alignment=1)
    ))

    doc.build(elementos)
    return response
